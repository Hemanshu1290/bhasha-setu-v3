"""
Bhasha Setu Backend v3
- Fixes the "yes plz apply" bug with a real pending_apply confirm step
- Optional AI-powered extraction + general scheme Q&A (Groq / OpenAI-compatible)
  Falls back to the deterministic regex extractor if no API key is set.

To enable AI mode, set an environment variable before starting the server:
  Windows (Command Prompt):  set GROQ_API_KEY=your_key_here
  Windows (PowerShell):      $env:GROQ_API_KEY="your_key_here"
  Mac/Linux:                 export GROQ_API_KEY=your_key_here

See README_AI.md for how to get a free Groq API key.
"""
from datetime import datetime, timezone
from pathlib import Path
import re
import os
import json
import urllib.request
import urllib.error
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from openpyxl import Workbook, load_workbook

DATA=Path(__file__).parent/'bhasha_setu_data.xlsx'
app=FastAPI(title='Bhasha Setu API')
app.add_middleware(CORSMiddleware,allow_origins=['*'],allow_methods=['*'],allow_headers=['*'])

# ---------- Optional AI config ----------
GROQ_API_KEY=os.environ.get('GROQ_API_KEY','').strip()
GROQ_MODEL=os.environ.get('GROQ_MODEL','llama-3.1-8b-instant')
GROQ_URL='https://api.groq.com/openai/v1/chat/completions'
AI_ENABLED=bool(GROQ_API_KEY)

class TextRequest(BaseModel):
    text: str

class ConfirmRequest(BaseModel):
    citizen_id: str
    updates: dict
    preview: str = ''

class ConfirmInsertRequest(BaseModel):
    new_citizen: dict
    preview: str = ''

class ConfirmApplyRequest(BaseModel):
    citizen_id: str
    scheme: str
    preview: str = ''

class SearchRequest(BaseModel):
    query: str

SHEETS={
    'Citizens':['Citizen_ID','Name','Mobile','State','District','City','Address','Has_Ration_Card'],
    'Scheme_Applications':['Application_ID','Citizen_ID','Scheme','Date','Status'],
    'Conversation_Log':['Timestamp','Citizen_ID','Language','Query','Intent'],
    'Updates_Log':['Timestamp','Citizen_ID','Field','Old Value','New Value']
}

SCHEMES_INFO={
    'PM Kisan':'₹6,000/year in 3 installments for eligible farmer families. Requires land records + Aadhaar-linked bank account.',
    'Ayushman Bharat':'Health cover of ₹5 lakh/family/year for secondary and tertiary hospitalization.',
    'PM Awas Yojana':'Financial assistance for building or purchasing a house for eligible low-income families.',
    'PM Ujjwala Yojana':'Free LPG gas connection for women from BPL households.',
}

def ensure_book():
    if DATA.exists(): 
        return
    wb=Workbook()
    wb.remove(wb.active)
    for n,h in SHEETS.items(): 
        ws=wb.create_sheet(n)
        ws.append(h)
        ws.freeze_panes='A2'
    wb['Citizens'].append(['C1024','Ravi Kumar','9876543210','Tamil Nadu','','Chennai','Chennai',True])
    wb['Scheme_Applications'].append(['A2051','C1024','PM Kisan',datetime.now().strftime('%Y-%m-%d'),'Pending'])
    wb.save(DATA)

def rows(sheet):
    ensure_book()
    wb=load_workbook(DATA)
    ws=wb[sheet]
    h=[c.value for c in ws[1]]
    return [dict(zip(h,r)) for r in ws.iter_rows(min_row=2,values_only=True) if any(v is not None for v in r)]

def match(e):
    for p in rows('Citizens'):
        if e.get('mobile') and str(p.get('Mobile'))==str(e['mobile']):
            return p
        if e.get('name') and str(p.get('Name','')).lower()==e['name'].lower():
            return p
    return None

def match_by_name_fuzzy(name):
    """Looser match used when a message just names someone, e.g. 'apply PM Kisan for Ravi Kumar'."""
    if not name:
        return None
    n=name.lower().strip()
    for p in rows('Citizens'):
        pname=str(p.get('Name','')).lower()
        if n==pname or n in pname or pname in n:
            return p
    return None

def next_citizen_id():
    ids=[p['Citizen_ID'] for p in rows('Citizens') if p.get('Citizen_ID')]
    nums=[int(re.sub(r'\D','',i) or 0) for i in ids]
    return f"C{(max(nums)+1) if nums else 1001}"

# ---------- Deterministic (regex) extractor — always available as fallback ----------
def extract_regex(text):
    t=text.lower()
    intent='GENERAL_QUERY'
    scheme=None
    for sname in SCHEMES_INFO:
        if sname.lower() in t or sname.lower().replace(' ','') in t.replace(' ',''):
            scheme=sname
            break
    if scheme is None and 'pm kisan' in t:
        scheme='PM Kisan'

    if any(x in t for x in ['apply','application','आवेदन','apply karna']):
        intent='NEW_APPLICATION'
    elif any(x in t for x in ['change','update','बदल','மாற்ற','tambaram','update karna']):
        intent='UPDATE_INFORMATION'
    elif any(x in t for x in ['status','स्थिति','நிலை','status check']):
        intent='APPLICATION_STATUS'
    elif any(x in t for x in ['how','information','जानकारी','எப்படி','kaise','inform']):
        intent='SCHEME_INFORMATION'
    elif any(x in t for x in ['search','find','खोज']):
        intent='SEARCH_CITIZEN'

    name=None; city=None; address=None; mobile=None

    m=re.search(r'(?:naam|name)\s+(?:hai\s+)?([A-Za-z][A-Za-z ]{1,40}?)(?:,|\s+hai|\s+main|\s+aur|$)',text,re.I)
    if m: name=m.group(1).strip()

    # "apply ... for <Name>" pattern — used when message doesn't say "my name"
    if not name:
        m=re.search(r'for\s+([A-Z][a-zA-Z]+(?:\s+[A-Z][a-zA-Z]+){0,2})',text)
        if m: name=m.group(1).strip()

    m=re.search(r'(?:mobile|फोन|number|नंबर|phone)[^0-9]{0,12}(\d{10})',text,re.I)
    if m: mobile=m.group(1)

    if 'tambaram' in t: address='Tambaram'

    m=re.search(r'(?:rehta|rehti|live|रहता|रहती|mein|में)\s+(?:hoon|हूँ|in)?\s*([A-Za-z ]{2,30}?)(?:\.|,|\s+aur|$)',text,re.I)
    if m: city=m.group(1).strip()

    ration=bool(re.search(r'ration card|राशन कार्ड|ரேஷன் கார்டு',t))

    e={k:v for k,v in {
        'name':name,'city':city,'address':address,'mobile':mobile,
        'has_ration_card':True if ration else None,'scheme':scheme
    }.items() if v is not None}

    lang='hi' if any(ord(c)>127 for c in text) or any(x in t for x in ['mera','mujhe','hai']) else 'en'

    return {'intent':intent,'language':lang,'entities':e,'raw_text':text}

# ---------- AI-powered extractor (optional, used when GROQ_API_KEY is set) ----------
def call_groq(messages, max_tokens=500, temperature=0.2):
    body=json.dumps({
        'model':GROQ_MODEL,
        'messages':messages,
        'max_tokens':max_tokens,
        'temperature':temperature
    }).encode('utf-8')
    req=urllib.request.Request(GROQ_URL,data=body,headers={
        'Content-Type':'application/json',
        'Authorization':f'Bearer {GROQ_API_KEY}'
    })
    try:
        with urllib.request.urlopen(req,timeout=15) as r:
            data=json.loads(r.read().decode('utf-8'))
            return data['choices'][0]['message']['content']
    except Exception as ex:
        return None

EXTRACT_SYSTEM_PROMPT="""You are an information-extraction engine for a citizen-services assistant.
Given a message in any language (Hindi, Tamil, Telugu, English, or Romanized Hindi), return ONLY a JSON object, no other text, with this exact shape:
{"intent":"NEW_APPLICATION|UPDATE_INFORMATION|APPLICATION_STATUS|SCHEME_INFORMATION|SEARCH_CITIZEN|GENERAL_QUERY",
 "language":"hi|ta|te|en",
 "entities":{"name":null,"city":null,"address":null,"mobile":null,"has_ration_card":null,"scheme":null},
 "raw_text":"<original text>"}
Rules:
- mobile must be exactly 10 digits or null.
- Only include a scheme if a specific government scheme is named (e.g. PM Kisan, Ayushman Bharat, PM Awas Yojana, PM Ujjwala Yojana).
- Omit fields you cannot find (use null).
- Return raw JSON only — no markdown fences, no commentary."""

def extract_ai(text):
    content=call_groq([
        {'role':'system','content':EXTRACT_SYSTEM_PROMPT},
        {'role':'user','content':text}
    ])
    if not content:
        return None
    try:
        cleaned=re.sub(r'^```json|```$','',content.strip(),flags=re.M).strip()
        parsed=json.loads(cleaned)
        parsed.setdefault('entities',{})
        parsed['entities']={k:v for k,v in parsed['entities'].items() if v not in (None,'','null')}
        parsed['raw_text']=text
        return parsed
    except Exception:
        return None

def extract(text):
    if AI_ENABLED:
        result=extract_ai(text)
        if result:
            return result
    return extract_regex(text)

def ai_scheme_answer(question):
    """General open-ended scheme Q&A via AI, used when AI_ENABLED and intent is GENERAL/SCHEME_INFORMATION."""
    if not AI_ENABLED:
        return None
    context='Known schemes: '+'; '.join(f'{k} — {v}' for k,v in SCHEMES_INFO.items())
    content=call_groq([
        {'role':'system','content':f'You are a helpful assistant for Indian citizens asking about government welfare schemes. Answer briefly (2-4 sentences) and in a friendly, clear tone. {context}. If asked about something you do not know, say so honestly rather than making it up.'},
        {'role':'user','content':question}
    ],max_tokens=300,temperature=0.4)
    return content

@app.on_event('startup')
def startup():
    ensure_book()

@app.get('/api/health')
def health():
    return {'ok':True,'ai_enabled':AI_ENABLED}

@app.get('/api/stats')
def stats():
    return {
        'citizens':len(rows('Citizens')),
        'applications':len(rows('Scheme_Applications')),
        'updates':len(rows('Updates_Log')),
        'ai_enabled':AI_ENABLED
    }

@app.get('/api/citizens')
def get_citizens():
    return {'data':rows('Citizens')}

@app.post('/api/search')
def search_citizen(req:SearchRequest):
    q=req.query.lower()
    results=[]
    for c in rows('Citizens'):
        if q in str(c.get('Name','')).lower() or q in str(c.get('Mobile','')):
            results.append(c)
    return {'results':results}

@app.post('/api/understand')
def understand(req:TextRequest):
    s=extract(req.text)
    e=s.get('entities',{})
    p=match(e)

    # If no direct match but a name-like phrase was found (e.g. "apply for Ravi Kumar"), try fuzzy match
    if not p and e.get('name'):
        p=match_by_name_fuzzy(e['name'])

    wb=load_workbook(DATA)
    wb['Conversation_Log'].append([
        datetime.now(timezone.utc).isoformat(),
        p.get('Citizen_ID') if p else '',
        s.get('language','en'),
        req.text,
        s.get('intent','GENERAL_QUERY')
    ])
    wb.save(DATA)

    intent=s.get('intent','GENERAL_QUERY')

    # UPDATE_INFORMATION: existing citizen update
    if intent=='UPDATE_INFORMATION' and p:
        u={k:v for k,v in e.items() if k in ['name','mobile','city','address','has_ration_card']}
        mapping={'name':'Name','mobile':'Mobile','city':'City','address':'Address','has_ration_card':'Has_Ration_Card'}
        if u:
            return {
                'structured':s,
                'pending_action':{
                    'citizen_id':p['Citizen_ID'],
                    'updates':u,
                    'preview':'; '.join(f"{k}: {p.get(mapping[k])} → {v}" for k,v in u.items())
                },
                'reply':''
            }

    # NEW_APPLICATION
    if intent=='NEW_APPLICATION':
        if p:
            scheme=e.get('scheme')
            if scheme:
                # Real fix: return a structured pending_apply instead of a dead-end text reply
                return {
                    'structured':s,
                    'pending_apply':{
                        'citizen_id':p['Citizen_ID'],
                        'scheme':scheme,
                        'preview':f"Apply for {scheme} on behalf of {p['Name']} ({p['Citizen_ID']})?"
                    },
                    'reply':''
                }
            return {'structured':s,'reply':f'Found existing citizen {p["Citizen_ID"]} ({p["Name"]}). Which scheme would you like to apply for?'}
        if e.get('name') and e.get('mobile'):
            new_citizen={
                'name':e['name'],'mobile':e['mobile'],'city':e.get('city',''),
                'address':e.get('address',e.get('city','')),
                'has_ration_card':e.get('has_ration_card',False),'scheme':e.get('scheme','')
            }
            preview=f"New citizen — {new_citizen['name']}, Mobile: {new_citizen['mobile']}, City: {new_citizen['city'] or '—'}"
            if new_citizen['scheme']:
                preview+=f", Apply for: {new_citizen['scheme']}"
            return {'structured':s,'pending_insert':{'new_citizen':new_citizen,'preview':preview},'reply':''}
        return {'structured':s,'reply':'I need at least a name and 10-digit mobile number to add a new citizen — or tell me an existing citizen\'s name to apply on their behalf.'}

    # SCHEME_INFORMATION
    if intent=='SCHEME_INFORMATION':
        ai_answer=ai_scheme_answer(req.text)
        if ai_answer:
            return {'structured':s,'reply':ai_answer}
        scheme=e.get('scheme')
        if scheme and scheme in SCHEMES_INFO:
            return {'structured':s,'reply':f"{scheme}: {SCHEMES_INFO[scheme]} No data was changed."}
        return {'structured':s,'reply':f"Here's what I know about government schemes: {'; '.join(SCHEMES_INFO.keys())}. Ask me about a specific one for details."}

    # APPLICATION_STATUS
    if intent=='APPLICATION_STATUS':
        if p:
            apps=[a for a in rows('Scheme_Applications') if a.get('Citizen_ID')==p['Citizen_ID']]
            if apps:
                status_txt='; '.join(f"{a.get('Scheme')}: {a.get('Status')}" for a in apps)
                return {'structured':s,'reply':f"Status for {p['Name']}: {status_txt}"}
            return {'structured':s,'reply':f"{p['Name']} has no applications yet."}
        return {'structured':s,'reply':'No citizen found to check status for. Include a name or mobile number.'}

    # SEARCH_CITIZEN
    if intent=='SEARCH_CITIZEN':
        q=(e.get('name') or '').lower()
        if q:
            matches=[c for c in rows('Citizens') if q in c.get('Name','').lower()]
            if matches:
                return {'structured':s,'reply':f"Found {len(matches)} citizen(s): {', '.join(c['Name'] for c in matches)}"}
        return {'structured':s,'reply':'No citizens found with that name. Try the Search Citizens tab.'}

    # GENERAL_QUERY — route to AI if available, for open-ended government scheme questions
    if AI_ENABLED:
        ai_answer=ai_scheme_answer(req.text)
        if ai_answer:
            return {'structured':s,'reply':ai_answer}

    return {'structured':s,'reply':'I understood your request. No spreadsheet change was made.'}

@app.post('/api/confirm')
def confirm(req:ConfirmRequest):
    wb=load_workbook(DATA)
    ws=wb['Citizens']
    h=[c.value for c in ws[1]]
    target=None
    for row in ws.iter_rows(min_row=2):
        if row[h.index('Citizen_ID')].value==req.citizen_id:
            target=row; break
    if not target:
        raise HTTPException(404,'Citizen record not found')
    mapping={'name':'Name','mobile':'Mobile','city':'City','address':'Address','has_ration_card':'Has_Ration_Card'}
    now=datetime.now(timezone.utc).isoformat()
    for k,new in req.updates.items():
        if k not in mapping: continue
        i=h.index(mapping[k]); old=target[i].value; target[i].value=new
        wb['Updates_Log'].append([now,req.citizen_id,mapping[k],old,new])
    wb.save(DATA)
    return {'reply':f'Updated {req.citizen_id} successfully. Changes logged.'}

@app.post('/api/confirm_insert')
def confirm_insert(req:ConfirmInsertRequest):
    wb=load_workbook(DATA)
    ws=wb['Citizens']
    nc=req.new_citizen
    cid=next_citizen_id()
    ws.append([cid,nc.get('name',''),nc.get('mobile',''),'','',nc.get('city',''),nc.get('address',''),bool(nc.get('has_ration_card',False))])
    if nc.get('scheme'):
        apps=wb['Scheme_Applications']
        aid=f"A{2000+len(rows('Scheme_Applications'))+1}"
        apps.append([aid,cid,nc['scheme'],datetime.now().strftime('%Y-%m-%d'),'Pending'])
    now=datetime.now(timezone.utc).isoformat()
    wb['Updates_Log'].append([now,cid,'NEW_CITIZEN','',f"Inserted {nc.get('name','')}"])
    wb.save(DATA)
    return {'reply':f'Added new citizen {cid} ({nc.get("name","")}) successfully.','citizen_id':cid}

@app.post('/api/confirm_apply')
def confirm_apply(req:ConfirmApplyRequest):
    """Fixes the original bug: applying a scheme for an already-known citizen now goes through
    a real confirm step instead of relying on free-text 'yes' with no memory."""
    wb=load_workbook(DATA)
    ws=wb['Citizens']
    h=[c.value for c in ws[1]]
    target=None
    name=''
    for row in ws.iter_rows(min_row=2):
        if row[h.index('Citizen_ID')].value==req.citizen_id:
            target=row
            name=row[h.index('Name')].value
            break
    if not target:
        raise HTTPException(404,'Citizen record not found')

    apps=wb['Scheme_Applications']
    aid=f"A{2000+len(rows('Scheme_Applications'))+1}"
    apps.append([aid,req.citizen_id,req.scheme,datetime.now().strftime('%Y-%m-%d'),'Submitted'])

    now=datetime.now(timezone.utc).isoformat()
    wb['Updates_Log'].append([now,req.citizen_id,'APPLICATION',f'Applied for {req.scheme}',''])
    wb.save(DATA)

    return {'reply':f'✓ Application created for {name}\nScheme: {req.scheme}\nCitizen ID: {req.citizen_id}\nStatus: Submitted','application_id':aid}

@app.get('/api/logs')
def get_logs():
    return {'updates':rows('Updates_Log'),'conversations':rows('Conversation_Log')}
